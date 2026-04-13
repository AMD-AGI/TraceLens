#!/usr/bin/env python3
###############################################################################
# Copyright (c) 2024 - 2025 Advanced Micro Devices, Inc. All rights reserved.
#
# See LICENSE for license information.
###############################################################################

"""
Generate a multi-sheet comparison report from two semantic breakdowns.

Combines TraceDiff outputs, comparison CSV, and a human-readable kernel mapping
into a single Excel workbook with per-sheet CSV exports.

Input:
  - Two semantic_labels.json files (one per trace)
  - comparison.csv (from match_and_compare.py)
  - TraceDiff output directory (from generate_semantic_diff.py)

Output:
  - comparison_report.xlsx (multi-sheet Excel workbook)
  - comparison_report_csvs/ (one CSV per sheet)

Sheets:
  1. kernel_mapping         - human-readable kernel-to-kernel mapping
  2. comparison             - block-level timing + roofline
  3. diff_stats             - TraceDiff-compatible per-kernel diff
  4. diff_stats_summary     - aggregated diff stats

Usage:
    python generate_comparison_report.py \\
        trace_a/semantic_labels.json trace_b/semantic_labels.json \\
        --comparison comparison.csv \\
        --diff-dir semantic_diff_output/ \\
        --name-a MI355 --name-b B200 \\
        [-o comparison_report.xlsx] \\
        [--output-csvs-dir comparison_report_csvs/]
"""

import argparse
import json
import math
import os
import sys
from collections import OrderedDict
from itertools import zip_longest

import pandas as pd
from openpyxl.styles import PatternFill

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "trace_breakdown"))
from category_mappings import get_perf_category


def load_labels(path):
    with open(path) as f:
        return json.load(f)


def _load_diff_stats(diff_dir, comparison_df):
    """Load diff_stats and summary from diff_dir, checking region subdirs."""
    diff_stats_path = os.path.join(diff_dir, "diff_stats.csv")
    summary_path = os.path.join(diff_dir, "diff_stats_unique_args_summary.csv")

    has_top_level = (
        os.path.exists(diff_stats_path) and os.path.getsize(diff_stats_path) > 10
    )
    if has_top_level:
        diff_df = pd.read_csv(diff_stats_path)
        sum_df = (
            pd.read_csv(summary_path)
            if os.path.exists(summary_path)
            else pd.DataFrame()
        )
        return diff_df, sum_df

    regions = (
        sorted(comparison_df["region"].unique())
        if "region" in comparison_df.columns
        else []
    )

    diff_parts = []
    summary_parts = []
    for region in regions:
        region_diff = os.path.join(diff_dir, region, "diff_stats.csv")
        region_sum = os.path.join(
            diff_dir, region, "diff_stats_unique_args_summary.csv"
        )
        if os.path.exists(region_diff) and os.path.getsize(region_diff) > 10:
            df = pd.read_csv(region_diff)
            df.insert(0, "region", region)
            diff_parts.append(df)
        if os.path.exists(region_sum) and os.path.getsize(region_sum) > 10:
            df = pd.read_csv(region_sum)
            df.insert(0, "region", region)
            summary_parts.append(df)

    diff_df = pd.concat(diff_parts, ignore_index=True) if diff_parts else pd.DataFrame()
    sum_df = (
        pd.concat(summary_parts, ignore_index=True) if summary_parts else pd.DataFrame()
    )
    return diff_df, sum_df


def build_kernel_mapping(labeled_a, labeled_b, name_a, name_b):
    """Build a human-readable kernel mapping DataFrame.

    Layout:
      semantic_block | {name_a} Kernels | {name_b} Kernels | perf_category
      QKV Projection | gemm_kernel_1    | cublas_gemm_1    | GEMM
                     | gemm_kernel_2    |                  |

    The semantic_block and perf_category are filled on the first row of each
    block and left blank on continuation rows. Kernel names are listed one per
    row (sorted). If one trace has more kernels in a block, the shorter side
    gets blank cells.
    """
    blocks_a = OrderedDict()
    block_cat = {}
    for k in labeled_a:
        b = k["semantic_block"]
        blocks_a.setdefault(b, []).append(k["name"])
        if b not in block_cat:
            block_cat[b] = k.get("perf_category") or get_perf_category(b)

    blocks_b = OrderedDict()
    for k in labeled_b:
        b = k["semantic_block"]
        blocks_b.setdefault(b, []).append(k["name"])
        if b not in block_cat:
            block_cat[b] = k.get("perf_category") or get_perf_category(b)

    all_blocks = list(
        OrderedDict.fromkeys(list(blocks_a.keys()) + list(blocks_b.keys()))
    )

    col_block = "semantic_block"
    col_a = f"{name_a} Kernels"
    col_b = f"{name_b} Kernels"
    col_cat = "perf_category"

    rows = []
    for block in all_blocks:
        kernels_a = sorted(set(blocks_a.get(block, [])))
        kernels_b = sorted(set(blocks_b.get(block, [])))

        paired = list(zip_longest(kernels_a, kernels_b, fillvalue=""))

        if not paired:
            paired = [("", "")]

        for i, (ka, kb) in enumerate(paired):
            is_first = i == 0
            rows.append(
                {
                    col_block: block if is_first else "",
                    col_a: ka,
                    col_b: kb,
                    col_cat: block_cat.get(block, "") if is_first else "",
                }
            )

    return pd.DataFrame(rows)


def _aggregate_by_category(df, name_a, name_b, gpu_scope_df=None):
    """Aggregate comparison rows by perf_category, appending a Total row.

    Returns a DataFrame with one row per perf_category plus a final Total row.
    Drops algorithm_order, avg_us, semantic_block, and semantic_group columns.
    """
    col_a_us = f"{name_a}_total_us"
    col_b_us = f"{name_b}_total_us"
    col_a_kcount = f"{name_a}_kernel_count"
    col_b_kcount = f"{name_b}_kernel_count"
    col_a_knames = f"{name_a}_kernel_names"
    col_b_knames = f"{name_b}_kernel_names"
    col_a_pct = f"{name_a}_pct"
    col_b_pct = f"{name_b}_pct"
    col_ratio = f"{name_a}_vs_{name_b}_ratio"
    col_gap = f"{name_a}_minus_{name_b}_us"
    gap_pct_col = f"{name_a} vs {name_b} %"

    for c in [
        col_a_us,
        col_b_us,
        col_a_kcount,
        col_b_kcount,
        col_a_pct,
        col_b_pct,
        col_gap,
    ]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    grouped = df.groupby("perf_category", sort=False)

    def concat_names(series):
        all_names = set()
        for val in series:
            if pd.notna(val) and val:
                all_names.update(n.strip() for n in str(val).split("|") if n.strip())
        return " | ".join(sorted(all_names)) if all_names else ""

    agg_rows = []
    for cat, grp in grouped:
        row = OrderedDict()
        row["perf_category"] = cat
        if col_a_knames in grp.columns:
            row[col_a_knames] = concat_names(grp[col_a_knames])
        row[col_a_kcount] = int(grp[col_a_kcount].sum())
        row[col_a_us] = round(grp[col_a_us].sum(), 2)
        row[col_a_pct] = round(grp[col_a_pct].sum(), 1)
        if col_b_knames in grp.columns:
            row[col_b_knames] = concat_names(grp[col_b_knames])
        row[col_b_kcount] = int(grp[col_b_kcount].sum())
        row[col_b_us] = round(grp[col_b_us].sum(), 2)
        row[col_b_pct] = round(grp[col_b_pct].sum(), 1)
        a_sum = grp[col_a_us].sum()
        b_sum = grp[col_b_us].sum()
        row[col_ratio] = round(a_sum / b_sum, 3) if b_sum > 0 else "inf"
        row[col_gap] = round(a_sum - b_sum, 2)
        if b_sum > 0:
            row[gap_pct_col] = round((a_sum - b_sum) / b_sum * 100, 1)
        else:
            row[gap_pct_col] = math.inf if a_sum > 0 else None
        agg_rows.append(row)

    result = pd.DataFrame(agg_rows)

    total_row = OrderedDict()
    total_row["perf_category"] = "Total"
    if col_a_knames in result.columns:
        total_row[col_a_knames] = ""
    total_row[col_a_kcount] = int(result[col_a_kcount].sum())
    total_row[col_a_us] = round(result[col_a_us].sum(), 2)
    total_row[col_a_pct] = round(result[col_a_pct].sum(), 1)
    if col_b_knames in result.columns:
        total_row[col_b_knames] = ""
    total_row[col_b_kcount] = int(result[col_b_kcount].sum())
    total_row[col_b_us] = round(result[col_b_us].sum(), 2)
    total_row[col_b_pct] = round(result[col_b_pct].sum(), 1)
    total_row[col_gap] = round(result[col_gap].sum(), 2)

    region_name = df["region"].iloc[0] if "region" in df.columns else None
    busy_a_col = f"{name_a}_busy_ms"
    busy_b_col = f"{name_b}_busy_ms"
    if (
        gpu_scope_df is not None
        and region_name
        and busy_a_col in gpu_scope_df.columns
        and busy_b_col in gpu_scope_df.columns
    ):
        region_row = gpu_scope_df[gpu_scope_df["region"] == region_name]
        if not region_row.empty:
            ba = float(region_row[busy_a_col].iloc[0])
            bb = float(region_row[busy_b_col].iloc[0])
            total_row[col_ratio] = round(ba / bb, 3) if bb > 0 else "inf"
        else:
            total_row[col_ratio] = ""
    else:
        total_a = result[col_a_us].sum()
        total_b = result[col_b_us].sum()
        total_row[col_ratio] = round(total_a / total_b, 3) if total_b > 0 else "inf"

    total_row[gap_pct_col] = ""
    result = pd.concat([result, pd.DataFrame([total_row])], ignore_index=True)
    return result


_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _apply_gap_fills(ws, gap_col_name):
    """Apply green/yellow/red fills to the gap percentage column."""
    header_row = [cell.value for cell in ws[1]]
    if gap_col_name not in header_row:
        return
    col_idx = header_row.index(gap_col_name) + 1

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        val = cell.value
        if val is None:
            continue
        try:
            val = float(val)
        except (TypeError, ValueError):
            continue
        if math.isinf(val):
            cell.fill = _FILL_RED
        elif val < 0:
            cell.fill = _FILL_GREEN
        elif val <= 10:
            cell.fill = _FILL_YELLOW
        else:
            cell.fill = _FILL_RED


def write_report(
    dict_name2df, output_xlsx_path=None, output_csvs_dir=None, gap_col=None
):
    """Write sheets as CSV files and/or a single Excel workbook."""
    if output_csvs_dir:
        os.makedirs(output_csvs_dir, exist_ok=True)
        for sheet_name, df in dict_name2df.items():
            csv_path = os.path.join(output_csvs_dir, f"{sheet_name}.csv")
            df.to_csv(csv_path, index=False)
        print(f"Wrote CSVs to {output_csvs_dir}/", file=sys.stderr)

    if output_xlsx_path:
        with pd.ExcelWriter(output_xlsx_path, engine="openpyxl") as writer:
            for sheet_name, df in dict_name2df.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            if gap_col:
                for sheet_name in writer.sheets:
                    if sheet_name.startswith("comparison"):
                        _apply_gap_fills(writer.sheets[sheet_name], gap_col)

        print(f"Wrote {output_xlsx_path}", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(
        description="Generate multi-sheet comparison report"
    )
    parser.add_argument("trace_a", help="Path to trace A semantic_labels.json")
    parser.add_argument("trace_b", help="Path to trace B semantic_labels.json")
    parser.add_argument(
        "--comparison",
        required=True,
        help="Path to comparison.csv (from match_and_compare.py)",
    )
    parser.add_argument(
        "--diff-dir",
        required=True,
        help="TraceDiff output directory (from generate_semantic_diff.py)",
    )
    parser.add_argument("--name-a", default="trace_a", help="Short name for trace A")
    parser.add_argument("--name-b", default="trace_b", help="Short name for trace B")
    parser.add_argument(
        "-o",
        "--output",
        default="comparison_report.xlsx",
        help="Output Excel path (default: comparison_report.xlsx)",
    )
    parser.add_argument(
        "--output-csvs-dir", default=None, help="Directory to write per-sheet CSV files"
    )
    args = parser.parse_args()

    data_a = load_labels(args.trace_a)
    data_b = load_labels(args.trace_b)
    labeled_a = data_a["labeled_kernels"]
    labeled_b = data_b["labeled_kernels"]

    print("Building kernel mapping...", file=sys.stderr)
    kernel_mapping_df = build_kernel_mapping(
        labeled_a, labeled_b, args.name_a, args.name_b
    )

    print("Loading comparison and diff data...", file=sys.stderr)
    comparison_df = pd.read_csv(args.comparison)

    diff_stats_df, diff_summary_df = _load_diff_stats(args.diff_dir, comparison_df)

    dict_name2df = OrderedDict()
    dict_name2df["kernel_mapping"] = kernel_mapping_df

    # Add gpu_scope_metrics sheet when comparison has gpu_timeline columns
    busy_a = f"{args.name_a}_busy_ms"
    idle_a = f"{args.name_a}_idle_pct"
    busy_b = f"{args.name_b}_busy_ms"
    idle_b = f"{args.name_b}_idle_pct"
    if all(c in comparison_df.columns for c in [busy_a, idle_a, busy_b, idle_b]):
        if "region" in comparison_df.columns:
            gpu_scope = (
                comparison_df.groupby("region")
                .first()[[busy_a, idle_a, busy_b, idle_b]]
                .reset_index()
            )
        else:
            gpu_scope = comparison_df[[busy_a, idle_a, busy_b, idle_b]].head(1)
            gpu_scope.insert(0, "region", "")
        dict_name2df["gpu_scope_metrics"] = gpu_scope
        comparison_df = comparison_df.drop(columns=[busy_a, idle_a, busy_b, idle_b])

    gap_col = f"{args.name_a} vs {args.name_b} %"
    gpu_scope_ref = dict_name2df.get("gpu_scope_metrics")

    if "region" in comparison_df.columns:
        agg_parts = []
        for region_name, region_df in comparison_df.groupby("region"):
            agg_region = _aggregate_by_category(
                region_df,
                args.name_a,
                args.name_b,
                gpu_scope_ref,
            )
            sheet_name = f"comparison_{region_name}"[:31]
            dict_name2df[sheet_name] = agg_region
            agg_with_region = agg_region.copy()
            agg_with_region.insert(0, "region", region_name)
            agg_parts.append(agg_with_region)
        dict_name2df["comparison"] = pd.concat(agg_parts, ignore_index=True)
    else:
        dict_name2df["comparison"] = _aggregate_by_category(
            comparison_df,
            args.name_a,
            args.name_b,
            gpu_scope_ref,
        )

    dict_name2df["diff_stats"] = diff_stats_df
    dict_name2df["diff_stats_summary"] = diff_summary_df

    output_xlsx = args.output
    output_csvs = args.output_csvs_dir

    if output_xlsx is None and output_csvs is None:
        output_xlsx = "comparison_report.xlsx"

    write_report(
        dict_name2df,
        output_xlsx_path=output_xlsx,
        output_csvs_dir=output_csvs,
        gap_col=gap_col if gap_col in comparison_df.columns else None,
    )

    for name, df in dict_name2df.items():
        print(f"  Sheet '{name}': {len(df)} rows", file=sys.stderr)


if __name__ == "__main__":
    main()
